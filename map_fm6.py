#!/usr/bin/env python3
"""Read Form FM6 Requirements.xlsx and create a JSON mapping for the GUI.

Usage: python3 map_fm6.py
This writes `fm6_mapping.json` in the project root.
Requires: openpyxl
"""
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    print('openpyxl is required. Install with: python3 -m pip install --user openpyxl')
    raise


def read_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows


def build_mapping(rows):
    # Heuristic: column C (index 2) contains field label; column D (index 3) contains Y/N for required.
    mapping = []
    for r in rows:
        if not any(r):
            continue
        # take label from col 3 if present, else col2 or col1
        label = None
        for ci in (2,1,0):
            if ci < len(r) and r[ci]:
                label = str(r[ci]).strip()
                break
        if not label:
            continue
        req = False
        if len(r) > 3 and r[3] is not None:
            v = str(r[3]).strip().upper()
            if v in ('Y','YES','REQUIRED','REQ'):
                req = True
        mapping.append({'key': sanitize_key(label), 'label': label, 'required': req})
    return mapping


def sanitize_key(label: str) -> str:
    k = label.lower()
    k = ''.join(ch if ch.isalnum() else '_' for ch in k)
    k = '_'.join(part for part in k.split('_') if part)
    return k[:64]


def main():
    p = Path('Form FM6 Requirements.xlsx')
    if not p.exists():
        print('Excel file not found:', p)
        sys.exit(1)
    rows = read_excel(p)
    mapping = build_mapping(rows)
    out = {'fields': mapping}
    with open('fm6_mapping.json','w',encoding='utf-8') as f:
        json.dump(out, f, indent=2)
    print('Wrote fm6_mapping.json with', len(mapping), 'fields')


if __name__ == '__main__':
    main()
